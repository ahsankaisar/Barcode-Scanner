from tkinter import *
import threading
import time
from sys import exit as Stp
from tkinter.ttk import *
from openpyxl import load_workbook
from openpyxl.cell import column_index_from_string as ColToInt
from time import strftime

date = strftime('%B %d %Y')  # Gets date in the from of MONTH DATE YEAR
saveFile = 'ATT.xlsx'  # Sets the save file
wb = load_workbook(saveFile)
ws = wb.active


def save():  # saves file
    wb.save(saveFile)


def find(findValue, row=None, column=None):
    if isinstance(findValue, str):
        findValue = findValue.lower()
    if row is None and column is None:
        for row in ws.iter_rows:
            for cell in row:
                if cell.value is findValue:
                    return cell.row, ColToInt(cell.column)

    elif row is not None and column is None:  # scanning through a row
        index = 1
        while True:
            cell = ws.cell(row=row, column=index)
            if cell.value is findValue:
                return cell.row, ColToInt(cell.column)
            index += 1

    elif row is None and column is not None:  # scanning through a column
        index = 2
        lastCell = ws.cell(row=index - 1, column=column)
        while True:
            cell = ws.cell(row=index, column=column)
            if lastCell.value is None and cell.value is None:
                return None, None
            if cell.value == findValue:
                return cell.row, ColToInt(cell.column)
            index += 1
            lastCell = cell


def create_attendance_date(row, column):  # Creates meeting date
    column = ColToInt(column)  # Converts letter column to number
    ws.cell(row=row, column=column, value=date)  # Adds date column
    save()  # Saves file


def find_empty_date_column():
    global attCol
    index = 2  # starts column count at 2('B')
    lastCell = ws['A1']  # first cell is set as 'A1'
    while True:  # runs loop to find empty cell in row
        cell = ws.cell(row=1, column=index)
        if lastCell.value is None and cell.value is None:
            # [date][None][None]...[None] finds 2 empty cells and the
            # uses the previous cell as the point at which to add the
            # date stamp
            create_attendance_date(lastCell.row, lastCell.column)
            attCol = (cell.column)
            break

        elif cell.value == date:

            # Makes attCol known to the whole program for attendance registration
            attCol = (cell.column)
            break

        # elif cell.value is None and lastCell.value is not None:
        index += 1
        lastCell = cell


def emptyRow():
    return find(None, None, 1)[0]


def number_check_in(num):
    timeIn = strftime("%I:%M %p")
    row, col = find(num, None, 1)
    if row is None and col is None:
        add_member()
    elif ws.cell(row=row, column=ColToInt(attCol)).value is None:
        ws.cell(row=row, column=ColToInt(attCol), value=timeIn)
    save()


def name_check_in(name):
    row, col = find(name, None, 2)
    if row is None and col is None:
        add_member()
    elif ws.cell(row=row, column=ColToInt(attCol)).value is None:
        timeIn = strftime("%I:%M %p")
        ws.cell(row=row, column=ColToInt(attCol), value=timeIn)
    save()


def check_in(identifier):
    find_empty_date_column()
    if isinstance(identifier, str):
        # Name based check in
        name_check_in(identifier.lower())
    else:
        # Number based check in
        number_check_in(identifier)


def approve_payment(identifier):
    if isinstance(identifier, str):
        identifier = identifier.lower()
        coll = 2
    else:
        coll = 1
    row, col = find(identifier, None, coll)
    if ws.cell(row=row, column=3).value is None:
        ws.cell(row=row, column=3, value='PAID')
    save()


def add_member():
    Fn = input('Enter First Name: ')
    Ln = input('Enter Last Name: ')
    FLname = Fn + Ln
    FLname = FLname.lower()
    number = int(input('Number: '))
    row = emptyRow()
    ws.cell(row=row, column=1, value=number)
    ws.cell(row=row, column=2, value=FLname)
    save()


def check_eligible(ID):
    if ID == 'None' or ID == '':
        return 'None'
    else:
        return 'Yes'


e = ['Wilson', 'Winters', 'Wise', 'Witt', 'Wolf', 'Wolfe', 'Wong', 'Wood', 'Woodard', 'Woods', 'Woodward', 'Wooten',
     'Workman', 'Wright', 'Wyatt', 'Wynn', 'Yang', 'Yates', 'York', 'Young', 'Zamora', 'Zimmerman']

root = Tk()
root.resizable(width=False, height=False)
Style().configure("TButton", relief="flat", padding=5, font='Times 14 bold')

menubar = Menu(root)
menuB = Menu(menubar, tearoff=0)
menuB2 = Menu(menubar, tearoff=0)
menuB3 = Menu(menubar, tearoff=1)
menuB.add_command(label="Exit", command=lambda: quit_handler())
menuB2.add_command(label="Attendance Graph")
menuB2.add_command(label='Total Member present')
menuB2.add_command(label='Credit Eligible Members')
menubar.add_cascade(label="File", menu=menuB)
menubar.add_cascade(label='Analyze', menu=menuB2)
menubar.add_cascade(label='Help', menu=menuB3)
root.configure(menu=menubar)

m1 = PanedWindow(height=650, width=1000, orient=VERTICAL)
m1.pack(fill=BOTH, expand=1)

# Barcode Entry
top = Frame(m1)
m1.add(top)

barcodeEntry = Entry(top, font='Times 24 bold', width=50)
barcodeEntry.pack(side=LEFT)


def check_entry_length():
    while True:
        lengthE = len(barcodeEntry.get())
        # print(lengthE)
        if lengthE == 9:
            retrieve_entry()
        time.sleep(0.09)


thr = threading.Thread(target=check_entry_length)
thr.daemon = True
thr.start()


def retrieve_entry():
    num = barcodeEntry.get()
    num = int(num)
    barcodeEntry.delete(0, END)
    check_in(num)


clearEntry = Button(top, text='Clear', command=lambda: barcodeEntry.delete(0, END)).pack(side=LEFT)

middle = Frame(m1)
m1.add(middle)

# Members List
scBr = Scrollbar(middle)
scBr.pack(side=RIGHT, fill=Y)

dataList = Listbox(middle, selectmode=SINGLE, font='Times 15 bold', yscrollcommand=scBr.set)
scBr.configure(command=dataList.yview)

dataList.insert(END, *e)
dataList.pack(fill=BOTH)

# Member information
bottom = Frame(m1)
m1.add(bottom)


def display_member_info(memberIndex):
    def member_info_template(name='Not Applicable', number='Not Applicable', paid='NO'):
        global a
        if name == paid == number and name == 'None':
            color1, color2, color3, color4 = None, None, None, None
        else:
            color1, color2, color3, color4 = 'white', 'lightgreen', 'lightgreen','red'
        eligible = check_eligible(memberIndex)

        if name == 'Not Applicable':
            color1 = 'Red'

        if paid == 'NO':
            color3 = 'Red'

        if number == 'Not Applicable':
            color2 = 'red'

        if eligible == 'Yes':
            color4 = 'lightgreen'

        nameLab = Label(bottom, text='Name: ' + name, font='Times 15 bold', background=color1)
        nameLab.grid(row=1, column=1, sticky=W)
        numLab = Label(bottom, text='Number: ' + str(number), font='Times 15 bold', background=color2)
        print(numLab)
        numLab.grid(row=2, column=1, sticky=W)
        payLab = Label(bottom, text='Paid: ' + paid, font='Times 15 bold', background=color3)
        payLab.grid(row=3, column=1, sticky=W)
        eligLab = Label(bottom, text='Eligible for Credit: ' + eligible, font='Times 15 bold', background=color4)
        eligLab.grid(row=4, column=1, sticky=W)


        return [nameLab, numLab, payLab,eligLab]

    memberIndex = (str(memberIndex))
    memberIndex = memberIndex.replace('(', '')
    memberIndex = memberIndex.replace(')', '')
    memberIndex = memberIndex.replace(',', '')
    if memberIndex is '':
        labelIDs = member_info_template('None', 'None', 'None')

    else:
        memberIndex = int(memberIndex)
        labelIDs = member_info_template(e[memberIndex])
    return labelIDs


def clean_mem_info_panel(labelIDList):
    if labelIDList is None:
        pass
    else:
        l1, l2, l3, l4 = labelIDList[0], labelIDList[1], labelIDList[2], labelIDList[3]
        l1.destroy()
        l2.destroy()
        l3.destroy()
        l4.destroy()


def check_selection():
    b = None
    lastPos = 0
    while True:
        curSelect = dataList.curselection()
        if curSelect != lastPos:
            clean_mem_info_panel(b)
            b = display_member_info(curSelect)
        else:
            pass
        time.sleep(0.1)
        lastPos = curSelect


listboxThr = threading.Thread(target=check_selection)
listboxThr.daemon = True
listboxThr.start()


def quit_handler():
    root.destroy()
    Stp()


barcodeEntry.focus_force()
root.protocol("WM_DELETE_WINDOW", quit_handler)
root.mainloop()
